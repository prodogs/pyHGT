from openpyxl import load_workbook


file_name = '/Users/frankfilippis/Library/CloudStorage/OneDrive-Personal/HG/HG_DATA.xlsx'
# Load workbook
wb = load_workbook(filename=file_name)

# Select a sheet
sheet = wb.active

# Print all values in the sheet
for row in sheet.iter_rows(values_only=True):
    print(row)